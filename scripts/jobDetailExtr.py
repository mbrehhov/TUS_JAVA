#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import load_workbook
import pandas
import re

call_details = dict()
counter = 0

f = open("/tmp/jobDetail.log", "w") 


book = openpyxl.load_workbook('/home/sagax/TUS/SEMESTER2/ENGPROJ/data/SampleDataEngineeringGroupSupervision.xlsm')

sheet = book.active

  

def printWriteOut(obj, obj2 = None):
    if obj2 == None:
        print(obj)
        f.write(obj+"\n")
    else:
        print(obj+":", obj2)
        f.write(obj+":"+str(obj2)+"\n")

def message(main):
    printWriteOut("MESSAGE",main)
    printWriteOut("***************************")
    
def contractor(message):
    # assumption is that contract data sequence is following
    list = []
    list.append('Job Type')
    list.append('Job Code')
    list.append('Due Date')
    list.append('Response Time')
    list.append('Pub Number')
    list.append('Pub Name')
    list.append('Pub Post Code')
    list.append('Pub Telephone Number')
    list.append('Placed by')
    list.append('Date placed')
    list.append('Trade Type')
    list.append('Sub Trade Type')
    list.append('Job description')
    for i in range(len(list)):
        if i < len(list) -1:
            regx = fr'{re.escape(list[i])}(.*?){re.escape(list[i+1])}'  #working 
            contr = re.findall(regx, message)
            if contr: 
                #possibleAnswer = btweenQuestions[0]
                printWriteOut(list[i],contr)
                


        else:
            print("contr:", list[i])
            regx = fr'{list[i]}(.*)'  #working 
            contr = re.findall(regx, message)
            if contr: 
                #possibleAnswer = btweenQuestions[0]
                printWriteOut(list[i],contr)
                
    printWriteOut("***************************")

def remString(main, sub):
    return main.replace(sub, ' ')
    
for row in range(1, 200):
    for col in sheet.iter_cols(3, 3):

        if (col[row].value==None):
            continue
        my_information = []

        #find all possible questions and store them into dict
        if "Job Type" in col[row].value:
            contractor(col[row].value)
            continue
        if (col[row].value).startswith("Additional Information:"):
            message(col[row].value)
            continue


        provide = re.findall(r'(Provide.*?[:?])(.*?(?=[A-Z]))', col[row].value)
        if provide:
            for provideRecord in provide:
                #printWriteOut("ProvideQuestion:")
                printWriteOut(provideRecord[0])
                
                col[row].value = remString(col[row].value,provideRecord[0])
                printWriteOut(provideRecord[1])
                
                col[row].value = remString(col[row].value,provideRecord[1])
 
        addInfo = re.findall(r'(Additional Information.*?[:?])(.*?(?=[.]))', col[row].value)
        if addInfo:
            for addInfoRecord in addInfo:
                printWriteOut(addInfoRecord[0])
                col[row].value = remString(col[row].value,addInfoRecord[0])
                printWriteOut(addInfoRecord[1])
                col[row].value = remString(col[row].value,addInfoRecord[1])



        brr = re.split(r'\?',col[row].value)
        counter = 1
        

        
        for element in brr :
            if len(brr)>1:
                ##assuming that last element in splitted array does not carry questionmark
                if (counter !=  len(brr) ) :
                    element=element+'?'
                counter+=1
            ## bit breaking regex
             
            if "H&S" in element:
                element= element.replace("H&S", "h&s")
                col[row].value = col[row].value.replace("H&S", "h&s")
            if "RAL" in element:
                element= element.replace("RAL", "ral")                
                col[row].value = col[row].value.replace("RAL", "ral")


            match = re.findall(r'[A-Z](?!.*[A-Z]).*\?', element)
            if match: 
                my_information.append(match[0])


              

        for i in range(len(my_information)):
            if i < len(my_information) -1:
                printWriteOut("question:", my_information[i])
                regx = fr'{re.escape(my_information[i])}(.*?){re.escape(my_information[i+1])}'  #working 
                
                btweenQuestions = re.findall(regx, (col[row].value))
                if btweenQuestions: 
                    
                    #print("possible answer:",btweenQuestions[0])
                    possibleAnswer = btweenQuestions[0]
                    
                                
                    
                    
                    ##
                    descr_regex = fr'(Describe.*?)(?=\:)(.*)?'   
                    descr = re.findall(descr_regex, btweenQuestions[0])
                    description = None
                    if descr:
                        description = descr[0][1]
                        #remove it fom answer
                        # ss = fr'(.*?){re.escape(descr[0][0])}' 
                        # result = re.search(btweenQuestions[0], ss)
                        # if(result):
                        answer = btweenQuestions[0].split(descr[0][0])
                        printWriteOut("answer:", answer[0])
                        printWriteOut(descr[0][0])
                        printWriteOut(descr[0][1])
                        
                    else:
                        #print("possible answer:",btweenQuestions[0])
                        printWriteOut("possible answer:",btweenQuestions[0])


                

                

            else:
                printWriteOut("possible answer:",btweenQuestions[0])
                regx = fr'{my_information[i]}(.*)'  #working 
                btweenQuestions = re.findall(regx, (col[row].value))
                if btweenQuestions: 
                    printWriteOut("possible answer:",btweenQuestions[0])
 

#Where is the door located?(.*?)How many doors ?
#: No Is this a fire door?(.*)
       # print(my_information)
        
        #Describe details details
        #    for describeR in describe:
        #        print("Describe:")
        #        print(provideRecord)

        

        printWriteOut("***************************")
       # print("> splitted array ")
       # print(brr)
        continue
        #retrive last question phrase
        #break it down
        
